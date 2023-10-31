# _____________________________________START________________________________________#

from openpyxl import Workbook, load_workbook
import subprocess as sb
import os 
import sys 
import keyboard as kb
import time


# lista workBook
wb_list = []
# script path
source_folder = sys.path[0]
# estrazione xlsx da directory append in wb_list
for (root,dir,file) in os.walk(source_folder):
    for f in file:
        if 'xlsx' in f:
           wb_list.append(f)



#iterating attraverso wb_list, scan ip address, ssh session 
    for i in wb_list:
        wb = load_workbook(i)
        sheet_names = wb.sheetnames
        print(f"Lista locations {sheet_names}")
        for i in sheet_names:
            ws = wb[f'{i}']
            for raw in range(2,2000):   
                value = ws[f'B{raw}'].value
                if value == None:
                    pass
                else:
                    with open('Output.txt', 'a') as Out:
                         # hold enter key
                        while True:
                            kb.send('enter')   
                            print('[+]-START SSH SESSION')
                            command = sb.run(['ssh', value], capture_output=True, text=True) 
                            
                            if command.stderr:
                                Out.write(f'Location {ws} {command.stderr}')
                            else:
                                pass
                            if command.stdout: 
                                Out.write(f'Location {ws} {command.stdout}')
                            else:
                                pass

                            print('\n')
                            print('Stop')
                            print('\n')
                            break
                
    input()





